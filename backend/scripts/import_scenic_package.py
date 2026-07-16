from __future__ import annotations

import argparse
from datetime import datetime
from decimal import Decimal
from hashlib import sha256
from pathlib import Path
import sys

from docx import Document
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.database import SessionLocal  # noqa: E402
from app.models.spot import ScenicSpot, SpotTag, VisitorBehaviorRecord  # noqa: E402


SPOT_DOCUMENT_NAME = "灵山胜境 景点结构化数据集.docx"
BEHAVIOR_WORKBOOK_NAME = "景点景区旅游数据行为分析数据.xlsx"
RELEVANT_ATTRACTIONS = {"灵山大佛", "灵山胜境", "禅意小镇·拈花湾"}
TAG_KEYWORDS = {
    "佛教文化": ("佛", "菩提", "坛城", "佛法"),
    "禅意": ("禅", "清净", "觉悟"),
    "历史文化": ("历史", "唐代", "宋代", "千年", "古刹"),
    "建筑艺术": ("建筑", "雕刻", "牌坊", "石柱", "浮雕", "塔"),
    "祈福": ("祈福", "朝拜", "朝圣", "吉祥"),
    "自然风光": ("太湖", "花海", "山林", "绿植", "湖", "谷"),
    "演艺": ("演出", "表演", "音乐", "动态景观"),
    "亲子": ("亲子", "孩童", "儿童"),
    "休闲": ("休憩", "休闲", "漫步", "步道"),
    "摄影": ("拍照", "打卡", "摄影"),
}


def compact_text(value: object) -> str:
    return " ".join(str(value or "").split())


def truncate(value: str, length: int) -> str:
    return value if len(value) <= length else value[: length - 1] + "…"


def derive_tags(text: str, scenic_area: str) -> list[str]:
    tags = [name for name, keywords in TAG_KEYWORDS.items() if any(keyword in text for keyword in keywords)]
    if scenic_area == "拈花湾禅意小镇" and "禅意" not in tags:
        tags.append("禅意")
    return tags[:20]


def derive_duration(description: str, highlights: str, opening_hours: str) -> int:
    text_length = len(description) + len(highlights)
    minutes = 15 + min(35, (text_length // 180) * 5)
    if "表演" in opening_hours or "演出" in opening_hours:
        minutes = max(minutes, 30)
    return max(15, min(60, minutes))


def import_spots(package_dir: Path) -> tuple[int, int]:
    source_path = package_dir / SPOT_DOCUMENT_NAME
    document = Document(source_path)
    created = 0
    updated = 0
    db = SessionLocal()
    try:
        for table in document.tables:
            headers = [compact_text(cell.text) for cell in table.rows[0].cells]
            for source_index, row in enumerate(table.rows[1:], start=1):
                values = [compact_text(cell.text) for cell in row.cells]
                record = dict(zip(headers, values, strict=True))
                external_id = record["景点ID"]
                if not external_id:
                    continue

                content = " ".join(values)
                summary_source = record["核心功能"] or record["游玩亮点"] or record["建筑/景观参数"]
                description = record["详细介绍"] or record["文化内涵"] or record["建筑/景观参数"]
                scenic_area = record["景区名称"]
                payload = {
                    "external_id": external_id,
                    "scenic_area": scenic_area,
                    "spot_type": "attraction",
                    "name": record["景点名称"],
                    "summary": truncate(summary_source, 255),
                    "description": description,
                    "location": truncate(record["具体位置"], 500) or None,
                    "opening_hours": truncate(record["演艺/开放信息"], 1000) or None,
                    "landscape_parameters": record["建筑/景观参数"] or None,
                    "cultural_context": record["文化内涵"] or None,
                    "highlights": record["游玩亮点"] or None,
                    "notes": record["备注"] or None,
                    "source_name": source_path.name,
                    "recommended_duration_minutes": derive_duration(
                        description, record["游玩亮点"], record["演艺/开放信息"]
                    ),
                    "priority": max(1, 101 - source_index),
                    "status": "enabled",
                    "cover_image_url": None,
                }
                spot = db.scalar(select(ScenicSpot).where(ScenicSpot.external_id == external_id))
                if spot is None:
                    spot = ScenicSpot(**payload)
                    db.add(spot)
                    db.flush()
                    created += 1
                else:
                    for key, value in payload.items():
                        setattr(spot, key, value)
                    spot.tags.clear()
                    db.flush()
                    updated += 1
                spot.tags.extend(SpotTag(name=tag) for tag in derive_tags(content, scenic_area))
        db.commit()
        return created, updated
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def decimal_value(value: object) -> Decimal:
    return Decimal(str(value)).quantize(Decimal("0.01"))


def record_key(values: list[object]) -> str:
    normalized = "|".join(value.isoformat() if isinstance(value, datetime) else str(value) for value in values)
    return sha256(normalized.encode("utf-8")).hexdigest()


def import_behavior(package_dir: Path, *, include_all: bool) -> tuple[int, int]:
    source_path = package_dir / BEHAVIOR_WORKBOOK_NAME
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    worksheet = workbook.active
    headers = [str(value) for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    inserted = 0
    skipped = 0
    batch: list[dict] = []
    db = SessionLocal()

    def flush_batch() -> None:
        nonlocal inserted
        if not batch:
            return
        statement = insert(VisitorBehaviorRecord).values(batch)
        statement = statement.on_conflict_do_nothing(index_elements=["source_record_key"])
        result = db.execute(statement)
        inserted += result.rowcount or 0
        db.commit()
        batch.clear()

    try:
        for values_tuple in worksheet.iter_rows(min_row=2, values_only=True):
            values = list(values_tuple)
            record = dict(zip(headers, values, strict=True))
            if not include_all and record["attraction_name"] not in RELEVANT_ATTRACTIONS:
                skipped += 1
                continue
            visit_date = record["visit_date"]
            if isinstance(visit_date, datetime):
                visit_date = visit_date.date()
            batch.append(
                {
                    "source_record_key": record_key(values),
                    "tourist_id": str(record["tourist_id"]),
                    "user_nickname": str(record["user_nickname"]),
                    "age": int(record["age"]),
                    "gender": str(record["gender"]),
                    "attraction_name": str(record["attraction_name"]),
                    "attraction_content": str(record["attraction_content"]),
                    "attraction_type": str(record["attraction_type"]),
                    "visit_date": visit_date,
                    "stay_duration_hours": decimal_value(record["stay_duration"]),
                    "ticket_cost": decimal_value(record["ticket_cost"]),
                    "food_cost": decimal_value(record["food_cost"]),
                    "shopping_cost": decimal_value(record["shopping_cost"]),
                    "transport_cost": decimal_value(record["transport_cost"]),
                    "entertainment_cost": decimal_value(record["entertainment_cost"]),
                    "total_cost": decimal_value(record["total_cost"]),
                    "group_size": int(record["group_size"]),
                    "satisfaction": int(record["satisfaction"]),
                    "source_name": source_path.name,
                }
            )
            if len(batch) >= 250:
                flush_batch()
        flush_batch()
        return inserted, skipped
    except Exception:
        db.rollback()
        raise
    finally:
        workbook.close()
        db.close()


def parse_args() -> argparse.Namespace:
    default_dir = Path(__file__).resolve().parents[2] / "示范景区资料包"
    parser = argparse.ArgumentParser(description="Import real scenic spot and visitor behavior data.")
    parser.add_argument("--package-dir", type=Path, default=default_dir)
    parser.add_argument("--include-all-behavior", action="store_true")
    parser.add_argument("--skip-spots", action="store_true")
    parser.add_argument("--skip-behavior", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    package_dir = args.package_dir.resolve()
    if not package_dir.is_dir():
        raise SystemExit(f"Package directory not found: {package_dir}")
    if not args.skip_spots:
        created, updated = import_spots(package_dir)
        print(f"Spots imported: created={created}, updated={updated}")
    if not args.skip_behavior:
        inserted, skipped = import_behavior(package_dir, include_all=args.include_all_behavior)
        print(f"Behavior imported: inserted={inserted}, skipped_unrelated={skipped}")


if __name__ == "__main__":
    main()
